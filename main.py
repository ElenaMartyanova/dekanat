import psycopg2
from psycopg2.extras import RealDictCursor
from datetime import datetime
from typing import Optional
from fastapi import FastAPI, Request, Form, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from pydantic import BaseModel, Field
from rapidfuzz import fuzz
from werkzeug.security import check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from io import BytesIO
import os
from dotenv import load_dotenv
from urllib.parse import quote
from starlette.exceptions import HTTPException as StarletteHTTPException
load_dotenv()

app = FastAPI(
    title="Virtual Dean",
    description="""
Информационный сервис виртуального деканата.

Система предоставляет веб-интерфейс и REST API для:
- поиска ответов по базе вопросов и ответов;
- получения списка категорий;
- отправки обратной связи;
- сохранения необработанных пользовательских вопросов;
- получения вопроса и ответа по идентификатору.

API может использоваться для интеграции с мобильным приложением, Telegram-ботом или внешними информационными системами.
""",
    version="1.0.0",
    openapi_tags=[
        {
            "name": "API",
            "description": "REST API для внешних клиентских приложений."
        },
        {
            "name": "Web",
            "description": "Пользовательский веб-интерфейс сервиса."
        },
        {
            "name": "Admin",
            "description": "Административная панель для управления системой."
        }
    ]
)

app.add_middleware(
    SessionMiddleware,
    secret_key=os.getenv("SECRET_KEY", "virtual_dean_secret_key")
)
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

@app.exception_handler(StarletteHTTPException)
async def custom_http_exception_handler(request: Request, exc: StarletteHTTPException):
    if exc.status_code == 404:
        language = request.query_params.get("lang") or request.session.get("lang") or "ru"
        language = normalize_language(language)

        return templates.TemplateResponse(
            request,
            "404.html",
            {
                "lang": language,
                "languages": LANGUAGES,
                "texts": TEXTS[language]
            },
            status_code=404
        )

    raise exc



DB_CONFIG = {
    "dbname": os.getenv("DB_NAME", "virtual_dean"),
    "user": os.getenv("DB_USER", "postgres"),
    "password": os.getenv("DB_PASSWORD", "admin123"),
    "host": os.getenv("DB_HOST", "localhost"),
    "port": os.getenv("DB_PORT", "5432")
}

LANGUAGES = {
    "ru": "RU",
    "en": "EN",
    "zh": "中文"
}


CATEGORY_NAMES = {
    "ru": {
        "migration": "Миграционный учёт",
        "medical": "Медицинское обслуживание",
        "housing": "Проживание",
        "documents": "Документы",
        "study": "Учебный процесс",
        "general": "Общие вопросы",
        "library": "Библиотека и учебные материалы",
        "military": "Воинский учет",
        "graduation": "Выпуск и ВКР",
        "activities": "Внеучебная деятельность",
        "main_answer": "Найденный ответ",
        "similar_questions": "Похожие вопросы",
        "international": "Иностранные студенты", 
        "it_support": "IT-поддержка и доступ к системам",      
        "transfer": "Перевод, отчисление, восстановление",
        "discipline": "Поведение и дисциплина",
        "practice": "Практика",
        "attestation": "Промежуточная аттестация",
        "psychological_service": "Психологическая служба",
        "disciplinary_commission": "Работа дисциплинарной комиссии",
        "scholarships": "Стипендии и материальная поддержка",
        "first_year": "Студенту первого курса",
        "employment": "Трудоустройство",
        "foreign_site_faq": "FAQ для иностранных граждан",
    },

    "en": {
        "migration": "Migration registration",
        "medical": "Medical services",
        "housing": "Accommodation",
        "documents": "Documents",
        "study": "Educational process",
        "general": "General questions",
        "library": "Library and study materials",
        "military": "Military registration",
        "graduation": "Graduation and final thesis",
        "activities": "Extracurricular activities",
        "main_answer": "Found answer",
        "similar_questions": "Similar questions",
        "international": "International students",
        "it_support": "IT support and system access",
        "transfer": "Transfer, expulsion and reinstatement",
        "discipline": "Conduct and discipline",
        "practice": "Internship",
        "attestation": "Interim assessment",
        "psychological_service": "Psychological support",
        "disciplinary_commission": "Disciplinary committee",
        "scholarships": "Scholarships and financial support",
        "first_year": "First-year students",
        "employment": "Employment",
        "foreign_site_faq": "FAQ for international students",

    },
    "zh": {
        "migration": "登记",
        "medical": "医疗服务",
        "housing": "住宿",
        "documents": "文件",
        "study": "学习过程",
        "general": "常见问题",
        "library": "图书馆和学习资料",
        "military": "兵役登记",
        "graduation": "毕业和毕业论文",
        "activities": "课外活动",
        "main_answer": "找到的答案",
        "similar_questions": "类似问题",
        "international": "国际学生",
        "it_support": "IT支持和系统访问",
        "transfer": "转学、退学和恢复学籍",
        "discipline": "行为与纪律",
        "practice": "实习实践",
        "attestation": "考试时期",
        "psychological_service": "心理疏导",
        "disciplinary_commission": "纪律委员会",
        "scholarships": "奖学金和经济资助",
        "first_year": "大一学生",
        "employment": "就业",
        "foreign_site_faq": "外国学生常见问题",

    }
}


TEXTS = {
    "ru": {
        "title": "Виртуальный деканат",
        "subtitle": "Информационная поддержка иностранных студентов",
        "search_placeholder": "Введите вопрос...",
        "search_button": "Найти",
        "categories": "Категории вопросов",
        "results": "Результаты поиска",
        "no_answer": "Точный ответ на данный вопрос пока не найден.",
        "similar": "Похожие вопросы или категории:",
        "contact_dean": "Рекомендуется обратиться в деканат для получения официальной консультации.",
        "admin_login": "Вход администратора",
        "username": "Логин",
        "password": "Пароль",
        "login": "Войти",
        "logout": "Выйти",
        "admin_panel": "Административная панель",
        "unanswered": "Необработанные запросы",
        "back": "Назад",
        "not_my_answer": "Нет ответа на мой вопрос",
        "query_saved": "Ваш запрос сохранён. При необходимости обратитесь в деканат.",
        "service_label": "Информационный сервис",
        "feedback_title": "Обратная связь",
        "feedback_placeholder": "Напишите вопрос, замечание или пожелание...",
        "feedback_rating": "Оцените работу сервиса",
        "feedback_send": "Отправить",
        "feedback_saved": "Спасибо! Ваше сообщение отправлено.",
        "feedback_admin": "Обратная связь пользователей",        


    },
    "en": {
        "title": "Virtual Dean's Office",
        "subtitle": "Information support for international students",
        "search_placeholder": "Enter your question...",
        "search_button": "Search",
        "categories": "Question categories",
        "results": "Search results",
        "no_answer": "An exact answer to this question has not been found yet.",
        "similar": "Similar questions or categories:",
        "contact_dean": "Please contact the dean's office for official consultation.",
        "admin_login": "Administrator login",
        "username": "Username",
        "password": "Password",
        "login": "Log in",
        "logout": "Log out",
        "admin_panel": "Admin panel",
        "unanswered": "Unanswered queries",
        "back": "Back",
        "not_my_answer": "My question was not answered",
        "query_saved": "Your query has been saved. If necessary, please contact the dean's office.",
        "service_label": "Information service",
        "feedback_title": "Feedback",
        "feedback_placeholder": "Write your question, comment, or suggestion...",
        "feedback_rating": "Rate the service",
        "feedback_send": "Send",
        "feedback_saved": "Thank you! Your message has been sent.",
        "feedback_admin": "User feedback",
    },
    "zh": {
        "title": "虚拟教务处",
        "subtitle": "为国际学生提供信息支持",
        "search_placeholder": "请输入问题...",
        "search_button": "搜索",
        "categories": "问题类别",
        "results": "搜索结果",
        "no_answer": "目前尚未找到该问题的准确答案。",
        "similar": "类似问题或类别：",
        "contact_dean": "建议联系教务处获取正式咨询。",
        "admin_login": "管理员登录",
        "username": "用户名",
        "password": "密码",
        "login": "登录",
        "logout": "退出",
        "admin_panel": "管理面板",
        "unanswered": "未处理的问题",
        "back": "返回",
        "not_my_answer": "没有找到我的问题答案",
        "query_saved": "您的问题已保存。如有需要，请联系教务处。",
        "service_label": "信息服务",
        "feedback_title": "反馈",
        "feedback_placeholder": "请写下您的问题、意见或建议...",
        "feedback_rating": "请评价服务",
        "feedback_send": "发送",
        "feedback_saved": "谢谢！您的反馈已提交。",
        "feedback_admin": "用户反馈",
    }
}

SEARCH_FILTERS = {
    "all": {
        "ru": "Все разделы",
        "en": "All sections",
        "zh": "全部类别",
        "categories": []
    },
    "medical_certificate": {
        "ru": "Справки по болезни",
        "en": "Medical certificates",
        "zh": "病假证明",
        "categories": ["medical", "study"]
    },
    "study_certificate": {
        "ru": "Справки об обучении",
        "en": "Study certificates",
        "zh": "在读证明",
        "categories": ["documents"]
    },
    "account": {
        "ru": "Личный кабинет и почта",
        "en": "Personal account and email",
        "zh": "个人账户和邮箱",
        "categories": ["it_support"]
    },
    "migration_visa": {
        "ru": "Виза и регистрация",
        "en": "Visa and registration",
        "zh": "签证和登记",
        "categories": [ "migration", "foreign_site_faq", "international"]
    },
    "housing": {
        "ru": "Общежитие",
        "en": "Dormitory",
        "zh": "宿舍",
        "categories": ["housing"]
    },
    "scholarships": {
        "ru": "Стипендии и материальная помощь",
        "en": "Scholarships and financial support",
        "zh": "奖学金和经济资助",
        "categories": ["scholarships"]
    },
    "study": {
        "ru": "Учебный процесс и экзамены",
        "en": "Study process and exams",
        "zh": "学习过程和考试",
        "categories": ["study", "attestation"]
    },
    "practice": {
        "ru": "Практика",
        "en": "Internship",
        "zh": "实习实践",
        "categories": ["practice"]
    },
    "employment": {
        "ru": "Трудоустройство",
        "en": "Employment",
        "zh": "就业",
        "categories": ["employment"]
    },
    "psychology": {
        "ru": "Психологическая служба",
        "en": "Psychological support",
        "zh": "心理疏导",
        "categories": ["psychological_service"]
    },
    "discipline": {
        "ru": "Дисциплина и комиссия",
        "en": "Discipline and committee",
        "zh": "纪律和委员会",
        "categories": ["discipline", "disciplinary_commission"]
    }
}

CATEGORY_ORDER = [
    # Блоки для иностранных студентов
    "international",
    "foreign_site_faq",
    "migration",

    # Учебные и основные студенческие вопросы
    "study",
    "general",
    "attestation",
    "housing",
    "documents",
    "it_support",
    "first_year",

    # Остальные разделы
    "medical",
    "practice",
    "employment",
    "scholarships",
    "transfer",
    "graduation",
    "library",
    "military",
    "activities",
    "discipline",
    "disciplinary_commission",
    "psychological_service",
]

FOREIGN_STUDENT_CATEGORIES = {
    "international",
    "foreign_site_faq",
    "migration",
    "general"
}

CATEGORY_GROUPS = {
    "ru": [
        {
            "title": "Часто задаваемые вопросы",
            "categories": ["international", "foreign_site_faq", "migration", "general"]
        },
        {
            "title": "Учёба и аттестация",
            "categories": ["study", "attestation", "practice", "graduation"]
        },
        {
            "title": "Стипендии, работа и дисциплина",
            "categories": [
                "scholarships",
                "employment",
                "transfer",
                "discipline",
                "disciplinary_commission",
                "library",
                "military",
                "activities"
            ]

        },
        {
            "title": "Проживание и поддержка",
            "categories": ["housing", "medical", "psychological_service"]
        },
        {
            "title": "Документы и цифровые сервисы",
            "categories": ["documents", "it_support", "first_year"]
        }
    ],
    "en": [
        {
            "title": "Frequently asked questions",
            "categories": ["international", "foreign_site_faq", "migration", "general"]
        },
        {
            "title": "Study and assessment",
            "categories": ["study", "attestation", "practice", "graduation"]
        },
        {
            "title": "Documents and digital services",
            "categories": ["documents", "it_support", "first_year"]
        },
        {
            "title": "Accommodation and support",
            "categories": ["housing", "medical", "psychological_service"]
        },
        {
            "title": "Scholarships, employment and discipline",
            "categories": [
                "scholarships",
                "employment",
                "transfer",
                "discipline",
                "disciplinary_commission",
                "library",
                "military",
                "activities"
            ]
        }
    ],
    "zh": [
        {
            "title": "常见问题",
            "categories": ["international", "foreign_site_faq", "migration", "general"]
        },
        {
            "title": "学习与考试",
            "categories": ["study", "attestation", "practice", "graduation"]
        },
        {
            "title": "文件和数字服务",
            "categories": ["documents", "it_support", "first_year"]
        },
        {
            "title": "住宿与支持",
            "categories": ["housing", "medical", "psychological_service"]
        },
        {
            "title": "奖学金、就业和纪律",
            "categories": [
                "scholarships",
                "employment",
                "transfer",
                "discipline",
                "disciplinary_commission",
                "library",
                "military",
                "activities"
            ]
        }
    ]
}
class DatabaseConnection:
    def __init__(self):
        self.connection = psycopg2.connect(
            **DB_CONFIG,
            cursor_factory=RealDictCursor
        )

    def execute(self, sql, params=None):
        cursor = self.connection.cursor()
        # Совместимость со старым SQLite-синтаксисом: ? -> %s для PostgreSQL
        sql = sql.replace("?", "%s")
        cursor.execute(sql, params or ())
        return cursor

    def commit(self):
        self.connection.commit()

    def close(self):
        self.connection.close()


def get_db_connection():
    return DatabaseConnection()



def get_categories(language):
    connection = get_db_connection()
    categories = connection.execute("SELECT * FROM Category").fetchall()
    connection.close()

    result = []

    for category in categories:
        code = category["code"]

        # Если категория есть в базе, но мы её не хотим показывать
        if code == "visa":
            continue

        result.append({
            "id": category["id"],
            "code": code,
            "name": CATEGORY_NAMES[language].get(code, code),
            "is_foreign": code in FOREIGN_STUDENT_CATEGORIES
        })

    result.sort(
        key=lambda item: CATEGORY_ORDER.index(item["code"])
        if item["code"] in CATEGORY_ORDER
        else 999
    )

    return result

def get_category_groups(language):
    categories = get_categories(language)
    categories_by_code = {
        category["code"]: category
        for category in categories
    }

    grouped_categories = []

    for group in CATEGORY_GROUPS[language]:
        items = []

        for category_code in group["categories"]:
            category = categories_by_code.get(category_code)

            if category:
                items.append(category)

        if items:
            grouped_categories.append({
                "title": group["title"],
                "categories": items
            })

    return grouped_categories

def get_search_filters(language):
    result = []

    for key, value in SEARCH_FILTERS.items():
        result.append({
            "code": key,
            "name": value.get(language, value["ru"])
        })

    return result



def normalize_query(query, language):
    query = query.lower().strip()

    replacements = {
        "ru": {
        
            # болезнь / справки / пропуски
            "куда нести справку": "болезнь медицинская справка справка по болезни деканат пропуск",
            "куда сдавать справку": "болезнь медицинская справка справка по болезни деканат пропуск",
            "куда отнести справку": "болезнь медицинская справка справка по болезни деканат пропуск",
            "нести справку": "болезнь медицинская справка справка по болезни деканат пропуск",
            "сдать справку": "болезнь медицинская справка справка по болезни деканат пропуск",
            "отнести справку": "болезнь медицинская справка справка по болезни деканат пропуск",

            "болел": "болезнь медицинская справка справка по болезни деканат пропуск",
            "болела": "болезнь медицинская справка справка по болезни деканат пропуск",
            "болею": "болезнь медицинская справка справка по болезни деканат пропуск",
            "заболел": "болезнь медицинская справка справка по болезни деканат пропуск",
            "заболела": "болезнь медицинская справка справка по болезни деканат пропуск",
            "больничный": "болезнь медицинская справка справка по болезни деканат пропуск",

            "справку": "справка",
            "справка": "справка",

            # личный кабинет / почта / доступ
            "забыл пароль": "забыл пароль восстановить пароль",
            "восстановить пароль": "забыл пароль восстановить пароль",
            "не помню пароль": "забыл пароль восстановить пароль",
            "не могу войти": "забыл пароль восстановить пароль личный кабинет вход ошибка поддержка",
            "не могу войти в личный кабинет": "забыл пароль восстановить пароль личный кабинет вход ошибка поддержка",
            "не заходит в личный кабинет": "забыл пароль восстановить пароль личный кабинет вход ошибка поддержка",
            "не получается войти в личный кабинет": "забыл пароль восстановить пароль личный кабинет вход ошибка поддержка",
            "проблема со входом": "забыл пароль восстановить пароль личный кабинет вход ошибка поддержка",
            "ошибка входа": "забыл пароль восстановить пароль личный кабинет вход ошибка поддержка",
            "лк": "личный кабинет",
            "кабинет": "личный кабинет",
            "пароль": "пароль",
            "почта": "корпоративная почта email доступ",
            # иностранные студенты / миграция / виза
            "как продлить визу": "продлить визу продление визы учебная виза",
            "продлить визу": "продлить визу продление визы учебная виза",
            "продление визы": "продлить визу продление визы учебная виза",
            "продлить учебную визу": "продлить визу продление визы учебная виза",
            "учебная виза": "учебная виза",
            "заканчивается виза": "продлить визу продление визы учебная виза",
            "виза заканчивается": "продлить визу продление визы учебная виза",
            "виза": "виза",
            "визу": "виза",

            "регистрация": "миграционный учет регистрация иностранный студент",
            "миграционка": "миграционная карта миграционный учет",
            "миграционная карта": "миграционная карта миграционный учет",
            "дмс": "медицинское страхование дмс страховка иностранный студент",
            "омс": "медицинское страхование омс иностранный студент",
            "страховка": "медицинское страхование дмс",

            # общежитие
            "общежитие": "общежитие проживание заселение",
            "общага": "общежитие проживание заселение",
            "заселиться": "общежитие заселение документы",
            "жить": "общежитие проживание",

            # учеба / экзамены / пересдачи
            "пересдача": "пересдача академическая задолженность экзамен",
            "пересдать": "пересдача академическая задолженность экзамен",
            "долг": "академическая задолженность пересдача",
            "долги": "академическая задолженность пересдача",
            "зачет": "зачет экзамен промежуточная аттестация",
            "зачёт": "зачет экзамен промежуточная аттестация",
            "экзамен": "экзамен сессия промежуточная аттестация",
            "сессия": "экзамен зачет промежуточная аттестация",
            "расписание": "расписание занятия пары",
            "пары": "занятия расписание",
            "пара": "занятие расписание",

            # работа / практика
            "работа": "трудоустройство вакансии работа",
            "работать": "трудоустройство вакансии работа иностранный студент",
            "вакансии": "трудоустройство вакансии ярмарка вакансий",
            "практика": "практика производственная практика учебная практика",

            # стипендии
            "стипендия": "стипендия материальная поддержка",
            "стипуха": "стипендия материальная поддержка",
            "матпомощь": "материальная помощь материальная поддержка",
            "материальная помощь": "материальная помощь материальная поддержка",

            # психолог
            "психолог": "психологическая служба психологическая помощь психологическая поддержка",
            "психологу": "психологическая служба психологическая помощь психологическая поддержка",

            # первый курс
            "первокурсник": "первый курс студент первого курса собрание факультета",
            "первый курс": "студент первого курса собрание факультета",
            "собрание": "собрание факультета первокурсник",
            "группа": "учебная группа список групп",

            # дисциплина / комиссия
            "выговор": "дисциплинарное взыскание дисциплинарная комиссия",
            "замечание": "дисциплинарное взыскание дисциплинарная комиссия",
            "нарушение": "дисциплина нарушение дисциплинарная комиссия",
            "комиссия": "дисциплинарная комиссия",

            # стоп-слова
            "куда": "",
            "что": "",
            "как": "",
            "где": "",
            "кто": "",
            "я": "",
            "мне": "",
            "надо": "",
            "нужно": "",
            "можно": "",
            "если": "",
            "это": ""
        },

        "en": {
            # medical certificates / illness
            "i was sick": "illness medical certificate absence dean office",
            "sick": "illness medical certificate dean office absence",
            "ill": "illness medical certificate dean office absence",
            "medical note": "medical certificate illness absence",
            "medical certificate": "medical certificate illness absence",
            "missed exam": "missed exam illness medical certificate",

            # account / password / email
            "forgot password": "forgot password reset password recover password",
            "forgot my password": "forgot password reset password recover password",
            "recover password": "forgot password reset password recover password",
            "reset password": "forgot password reset password recover password",
            "password reset": "forgot password reset password recover password",
            "lost password": "forgot password reset password recover password",

            "can't log in": "personal account login access password",
            "cannot log in": "personal account login access password",
            "personal account": "personal account student account",
            "student account": "personal account student account",
            "password": "password",
            "login": "personal account login access",
            "email": "corporate email access",
            "corporate email": "corporate email access",

            # visa / registration / insurance
            "visa expires": "visa extension international student",
            "extend visa": "visa extension international student",
            "visa": "visa extension international student",
            "migration registration": "migration registration international student",
            "registration": "migration registration international student",
            "migration card": "migration card migration registration",
            "lost migration card": "lost migration card migration registration",
            "insurance": "medical insurance dms oms",
            "medical insurance": "medical insurance dms oms",
            "dms": "medical insurance dms",
            "oms": "medical insurance oms",

            # dormitory
            "dorm": "dormitory accommodation housing",
            "dormitory": "dormitory accommodation housing",
            "hostel": "dormitory accommodation housing",
            "move into dormitory": "dormitory accommodation housing documents",
            "under 18": "dormitory under 18 parents guardian",

            # study / exams
            "retake": "exam retake academic debt",
            "exam": "exam interim assessment",
            "grades": "grades personal account",
            "grades are not shown": "grades personal account support",
            "schedule": "schedule classes timetable",

            # employment / scholarship / psychology
            "work": "employment vacancies job",
            "job": "employment vacancies job",
            "vacancies": "employment vacancies job",
            "scholarship": "scholarship financial support",
            "financial support": "scholarship financial support",
            "psychologist": "psychological support service",
            "psychological support": "psychological support service",

            # stop words
            "where": "",
            "how": "",
            "what": "",
            "need": "",
            "can": "",
            "should": "",
            "my": "",
            "i": "",
            "where to submit certificate": "medical certificate dean office illness absence",
            "submit medical certificate": "medical certificate dean office illness absence",
            "illness certificate": "medical certificate illness absence",

            "visa is expiring": "visa extension international student",
            "my visa expires": "visa extension international student",
            "extend my visa": "visa extension international student",

            "where is dormitory": "dormitory accommodation address",
            "dormitory documents": "dormitory accommodation documents",
            "move in": "dormitory accommodation housing",

            "first year": "first year students faculty meeting classes",
            "faculty meeting": "first year students faculty meeting",

            "academic debt": "academic debt retake exam",
            "missed test": "missed exam retake medical certificate",
            "retake exam": "exam retake academic debt",

            "study certificate": "study certificate documents dean office",
            "certificate of study": "study certificate documents dean office",
        },

        "zh": {
            # болезнь / справки
            "我生病了": "生病 医疗证明 院系办公室 缺课",
            "生病": "医疗证明 院系办公室 缺课",
            "病假": "医疗证明 院系办公室 缺课",
            "病假证明": "医疗证明 院系办公室 缺课",
            "医疗证明": "医疗证明 院系办公室 缺课",
            "没参加考试": "生病 医疗证明 考试 缺课",

            # личный кабинет / пароль / почта
            "忘记密码": "忘记密码 恢复密码 个人账户 登录",
            "密码忘了": "忘记密码 恢复密码 个人账户 登录",
            "恢复密码": "忘记密码 恢复密码 个人账户 登录",
            "无法登录": "个人账户 登录 密码",
            "不能登录": "个人账户 登录 密码",
            "个人账户": "个人账户 登录",
            "学生账户": "个人账户 登录",
            "密码": "密码",
            "登录": "个人账户 登录",
            "邮箱": "学校邮箱 企业邮箱",
            "学校邮箱": "学校邮箱 企业邮箱",

            # виза / регистрация / страховка
            "签证快到期了": "签证 延期 外国学生",
            "签证延期": "签证 延期 外国学生",
            "延期签证": "签证 延期 外国学生",
            "签证": "签证 延期 外国学生",
            "居留登记": "居留登记 外国学生",
            "登记": "居留登记 外国学生",
            "移民卡": "移民卡 居留登记",
            "移民卡丢了": "移民卡 丢失 居留登记",
            "保险": "医疗保险 ДМС ОМС",
            "医疗保险": "医疗保险 ДМС ОМС",
            "ДМС": "医疗保险 ДМС",
            "ОМС": "医疗保险 ОМС",

            # общежитие
            "宿舍": "宿舍 住宿 入住",
            "住宿": "宿舍 住宿 入住",
            "入住宿舍": "宿舍 住宿 入住 文件",
            "宿舍文件": "宿舍 住宿 文件",
            "没满18岁": "宿舍 未满18岁 父母 监护人",

            # учеба / экзамены
            "补考": "补考 考试 欠考的课程",
            "考试": "考试 考试时期",
            "成绩": "成绩 个人账户",
            "成绩没有显示": "成绩 个人账户",
            "课表": "课表 课程表",
            "课程表": "课表 课程表",

            # работа / стипендия / психолог
            "工作": "就业 招聘 工作",
            "就业": "就业 招聘 工作",
            "招聘": "就业 招聘 工作",
            "奖学金": "奖学金 经济资助",
            "经济资助": "奖学金 经济资助",
            "心理": "心理疏导 心理支持",
            "心理疏导": "心理疏导 心理支持",
            "证明交到哪里": "医疗证明 院系办公室 缺课",
            "病假证明交到哪里": "医疗证明 院系办公室 缺课",
            "生病证明": "医疗证明 院系办公室 缺课",

            "签证快过期了": "签证 延期 外国学生",
            "签证到期": "签证 延期 外国学生",
            "怎么延期签证": "签证 延期 外国学生",

            "宿舍在哪里": "宿舍 住宿 地址",
            "宿舍需要哪些文件": "宿舍 住宿 文件",
            "怎么入住宿舍": "宿舍 住宿 入住",

            "大一": "大一学生 学院大会 开学",
            "新生": "大一学生 学院大会 开学",
            "学院大会": "大一学生 学院大会",

            "欠考": "补考 考试 欠考的课程",
            "重考": "补考 考试 欠考的课程",
            "没参加考试": "生病 医疗证明 考试 缺课",

            "在读证明": "在读证明 文件 院系办公室",
            "学习证明": "在读证明 文件 院系办公室",
        }
    }

    for old, new in replacements.get(language, {}).items():
        query = query.replace(old, new)

    words = query.split()
    words = [word.strip(".,!?;:()[]{}«»\"'") for word in words]
    words = [word for word in words if len(word) > 1]

    return words


def is_medical_certificate_query(query):
    query = query.lower().strip()

    phrases = [
        "куда нести справку",
        "куда сдавать справку",
        "куда отнести справку",
        "нести справку",
        "сдать справку",
        "отнести справку",
        "справка по болезни",
        "болел",
        "болела",
        "болею",
        "заболел",
        "заболела",
        "больничный"
    ]

    return any(phrase in query for phrase in phrases)

def detect_search_filter(query, language):
    query = query.lower().strip()

    rules = {
        "ru": {
            "migration_visa": [
                "виза",
                "визу",
                "визы",
                "визе",
                "визовый",
                "визовая",
                "регистрация",
                "миграционный",
                "миграционная карта",
                "миграционка",
                "рвпо",
                "дмс",
                "омс"
            ],
            "medical_certificate": [
                "справка",
                "болел",
                "болела",
                "болею",
                "заболел",
                "заболела",
                "больничный",
                "болезнь"
            ],
            "account": [
                "личный кабинет",
                "лк",
                "пароль",
                "почта",
                "не могу войти",
                "не получается войти",
                "забыл пароль",
                "восстановить пароль"
            ],
            "housing": [
                "общежитие",
                "общага",
                "заселиться",
                "проживание"
            ],
            "scholarships": [
                "стипендия",
                "стипуха",
                "матпомощь",
                "материальная помощь"
            ],
            "practice": [
                "практика"
            ],
            "employment": [
                "работа",
                "работать",
                "вакансии",
                "трудоустройство"
            ],
            "study": [
                "пересдача",
                "экзамен",
                "зачет",
                "зачёт",
                "сессия",
                "расписание",
                "академическая задолженность"
            ]
        },
        "en": {
            "migration_visa": [
                "visa",
                "registration",
                "migration",
                "migration card",
                "insurance",
                "dms",
                "oms"
            ],
            "medical_certificate": [
                "medical certificate",
                "sick",
                "ill",
                "illness",
                "medical note"
            ],
            "account": [
                "personal account",
                "student account",
                "password",
                "login",
                "email",
                "forgot password",
                "reset password"
            ],
            "housing": [
                "dormitory",
                "dorm",
                "hostel",
                "accommodation"
            ],
            "scholarships": [
                "scholarship",
                "financial support"
            ],
            "practice": [
                "internship",
                "practice"
            ],
            "employment": [
                "work",
                "job",
                "vacancies",
                "employment"
            ],
            "study": [
                "exam",
                "retake",
                "schedule",
                "academic debt"
            ]
        },
        "zh": {
            "migration_visa": [
                "签证",
                "登记",
                "居留登记",
                "移民卡",
                "保险"
            ],
            "medical_certificate": [
                "生病",
                "病假",
                "医疗证明",
                "病假证明"
            ],
            "account": [
                "密码",
                "登录",
                "个人账户",
                "学生账户",
                "邮箱",
                "忘记密码"
            ],
            "housing": [
                "宿舍",
                "住宿"
            ],
            "scholarships": [
                "奖学金",
                "经济资助"
            ],
            "practice": [
                "实习"
            ],
            "employment": [
                "工作",
                "就业",
                "招聘"
            ],
            "study": [
                "考试",
                "补考",
                "课表",
                "课程表"
            ]
        }
    }

    language_rules = rules.get(language, rules["ru"])

    for filter_code, keywords in language_rules.items():
        if any(keyword in query for keyword in keywords):
            return filter_code

    return "all"

def search_answers(query, language, search_filter="all"):
    words = normalize_query(query, language)
    original_query = query.lower().strip()
    normalized_query = " ".join(words)

    if search_filter == "all":
        detected_filter = detect_search_filter(query, language)
    else:
        detected_filter = search_filter

    selected_categories = SEARCH_FILTERS.get(
        detected_filter,
        SEARCH_FILTERS["all"]
    )["categories"]

    connection = get_db_connection()
    records = connection.execute("""
        SELECT 
            Translation.id,
            Translation.question_id,
            Translation.question_text,
            Translation.answer_text,
            Category.code AS category_code
        FROM Translation
        JOIN Question ON Translation.question_id = Question.id
        JOIN Category ON Question.category_id = Category.id
        WHERE Translation.language = ?
    """, (language,)).fetchall()
    connection.close()

    results = []

    for record in records:
        question_text = record["question_text"].lower()
        answer_text = record["answer_text"].lower()
        category_code = record["category_code"].lower()

        # Если пользователь выбрал фильтр, ищем только в нужных категориях.
        # Это не даёт нерелевантным разделам перебивать правильные ответы.
        if selected_categories and category_code not in selected_categories:
            continue

        score = 0

        # 1. Сравнение исходного запроса с вопросом.
        # token_set_ratio хорошо работает, когда слова в разном порядке
        # или в вопросе есть дополнительные слова.
        score += fuzz.token_set_ratio(original_query, question_text) * 0.45

        # 2. Частичное совпадение исходного запроса с вопросом.
        # Помогает при коротких запросах: "забыл пароль", "куда справку".
        score += fuzz.partial_ratio(original_query, question_text) * 0.25

        # 3. Сравнение нормализованного запроса с вопросом.
        # Тут учитываются наши сокращения и синонимы: "лк" -> "личный кабинет".
        if normalized_query:
            score += fuzz.token_set_ratio(normalized_query, question_text) * 0.35

        # 4. Ответ тоже учитываем, но слабее,
        # чтобы ответ не перебивал более подходящий вопрос.
        score += fuzz.token_set_ratio(original_query, answer_text) * 0.08

        if normalized_query:
            score += fuzz.token_set_ratio(normalized_query, answer_text) * 0.08

        # 5. Точные совпадения — сильный бонус.
        if original_query and original_query in question_text:
            score += 35

        if normalized_query and normalized_query in question_text:
            score += 25

        # 6. Совпадение отдельных слов в вопросе.
        # Вопрос важнее ответа.
        for word in words:
            if word in question_text:
                score += 8
            elif word in answer_text:
                score += 2

        # 7. Небольшая поддержка выбранного фильтра уже есть через ограничение,
        # но для "Все разделы" ничего не добавляем.
        if selected_categories and category_code in selected_categories:
            score += 5

        # 8. Отсекаем слабые совпадения.
        if score >= 35:
            results.append({
                "question_text": record["question_text"],
                "answer_text": record["answer_text"],
                "category_code": record["category_code"],
                "score": round(score, 2)
            })

    results.sort(key=lambda item: item["score"], reverse=True)

    main_result = results[0] if results else None
    similar_results = results[1:6] if len(results) > 1 else []

    return main_result, similar_results

def save_unanswered_query(query, language):
    connection = get_db_connection()
    connection.execute("""
        INSERT INTO UnansweredQuery (query_text, language, created_at, status)
        VALUES (?, ?, ?, ?)
    """, (query, language, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "new"))
    connection.commit()
    connection.close()

def save_feedback(query_text, feedback_text, rating, language):
    connection = get_db_connection()
    connection.execute("""
        INSERT INTO Feedback (query_text, feedback_text, rating, language, created_at, status)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (
        query_text,
        feedback_text,
        rating,
        language,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "new"
    ))
    connection.commit()
    connection.close()



def detect_text_language(text: str) -> str:
    text = text.strip()

    if not text:
        return "empty"

    has_cyrillic = any("а" <= char.lower() <= "я" or char.lower() == "ё" for char in text)
    has_chinese = any("\u4e00" <= char <= "\u9fff" for char in text)
    has_latin = any("a" <= char.lower() <= "z" for char in text)

    if has_chinese:
        return "zh"

    if has_cyrillic:
        return "ru"

    if has_latin:
        return "en"

    return "unknown"


def validate_translation_language(text: str, expected_language: str) -> bool:
    text = text.strip()

    if not text:
        return True

    detected_language = detect_text_language(text)

    if detected_language == "unknown":
        return True

    return detected_language == expected_language

def get_search_language_error(query: str, selected_language: str):
    detected_language = detect_text_language(query)

    if detected_language in ["empty", "unknown"]:
        return None

    if detected_language == selected_language:
        return None

    messages = {
        "ru": {
            "main": "Вы выбрали русский язык, но вопрос написан на другом языке. Переключите язык сайта или задайте вопрос на русском языке.",
            "small": [
                "You selected Russian, but the question is written in another language. Please switch the website language or ask the question in Russian.",
                "您选择了俄语，但问题是用其他语言写的。请切换网站语言，或用俄语提问。"
            ]
        },
        "en": {
            "main": "You selected English, but the question is written in another language. Please switch the website language or ask the question in English.",
            "small": [
                "Вы выбрали английский язык, но вопрос написан на другом языке. Переключите язык сайта или задайте вопрос на английском языке.",
                "您选择了英语，但问题是用其他语言写的。请切换网站语言，或用英语提问。"
            ]
        },
        "zh": {
            "main": "您选择了中文，但问题是用其他语言写的。请切换网站语言，或使用中文提问。",
            "small": [
                "Вы выбрали китайский язык, но вопрос написан на другом языке. Переключите язык сайта или задайте вопрос на китайском языке.",
                "You selected Chinese, but the question is written in another language. Please switch the website language or ask the question in Chinese."
            ]
        }
    }

    return messages.get(selected_language, messages["ru"])

def validate_question_languages(translations):
    errors = []

    for language_code, question_text, answer_text in translations:
        if question_text.strip() and not validate_translation_language(question_text, language_code):
            errors.append(f"Вопрос для языка {language_code.upper()} заполнен не на том языке.")

        if answer_text.strip() and not validate_translation_language(answer_text, language_code):
            errors.append(f"Ответ для языка {language_code.upper()} заполнен не на том языке.")

    return errors

def normalize_language(language: str) -> str:
    if language not in LANGUAGES:
        return "ru"
    return language


def get_request_language(request: Request) -> str:
    language = request.query_params.get("lang") or request.session.get("lang") or "ru"
    language = normalize_language(language)
    request.session["lang"] = language
    return language


def is_admin_logged_in(request: Request) -> bool:
    return bool(request.session.get("admin_logged_in"))


def redirect_to(url: str) -> RedirectResponse:
    return RedirectResponse(url=url, status_code=303)

def render_page(request: Request, template_name: str, context: dict):
    return templates.TemplateResponse(
        request,
        template_name,
        context
    )

def format_excel_workbook(workbook):
    header_fill = PatternFill("solid", fgColor="372579")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column_cells in sheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            width = min(max(max_length + 2, 12), 45)
            sheet.column_dimensions[column_letter].width = width

        sheet.freeze_panes = "A2"

class SearchRequest(BaseModel):
    query: str = Field(
        ...,
        min_length=1,
        examples=["забыл пароль"],
        description="Текст вопроса пользователя"
    )
    lang: str = Field(
        "ru",
        examples=["ru"],
        description="Язык запроса: ru, en или zh"
    )
    search_filter: str = Field(
        "all",
        examples=["account"],
        description="Код фильтра поиска. Например: all, account, migration_visa, housing"
    )


class FeedbackRequest(BaseModel):
    query: Optional[str] = Field(
        "",
        examples=["забыл пароль"],
        description="Вопрос пользователя, к которому относится обратная связь"
    )
    feedback_text: Optional[str] = Field(
        "",
        examples=["Ответ помог, всё понятно"],
        description="Текст обратной связи"
    )
    rating: Optional[int] = Field(
        None,
        ge=1,
        le=5,
        examples=[5],
        description="Оценка работы сервиса от 1 до 5"
    )
    lang: str = Field(
        "ru",
        examples=["ru"],
        description="Язык сообщения: ru, en или zh"
    )


class UnansweredRequest(BaseModel):
    query: str = Field(
        ...,
        min_length=1,
        examples=["где получить новую карту студента"],
        description="Вопрос пользователя, на который не был найден ответ"
    )
    lang: str = Field(
        "ru",
        examples=["ru"],
        description="Язык вопроса: ru, en или zh"
    )


@app.get(
    "/api/categories",
    tags=["API"],
    summary="Получить список категорий",
    description="Возвращает список категорий вопросов на выбранном языке."
)
def api_categories(lang: str = "ru"):
    language = normalize_language(lang)
    return {
        "status": "success",
        "language": language,
        "categories": get_categories(language)
    }

@app.get("/search", tags=["Web"])
def search_get(request: Request):
    language = get_request_language(request)
    return redirect_to(f"/?lang={language}")

@app.post(
    "/api/search",
    tags=["API"],
    summary="Найти ответ на вопрос",
    description="Выполняет поиск по базе вопросов и ответов с учётом языка и выбранного фильтра."
)
def api_search(data: SearchRequest):
    language = normalize_language(data.lang)
    query = data.query.strip()

    if not query:
        raise HTTPException(status_code=400, detail="Query is required")

    main_result, similar_results = search_answers(query, language, data.search_filter)

    if not main_result:
        save_unanswered_query(query, language)

    return {
        "status": "success",
        "language": language,
        "query": query,
        "search_filter": data.search_filter,
        "main_result": main_result,
        "similar_results": similar_results
    }


@app.post(
    "/api/feedback",
    tags=["API"],
    summary="Отправить обратную связь",
    description="Сохраняет отзыв пользователя и оценку качества работы сервиса."
)
def api_feedback(data: FeedbackRequest):
    language = normalize_language(data.lang)
    query = data.query.strip() if data.query else ""
    feedback_text = data.feedback_text.strip() if data.feedback_text else ""

    if not feedback_text and data.rating is None:
        raise HTTPException(status_code=400, detail="Feedback text or rating is required")

    save_feedback(query, feedback_text, data.rating, language)
    return {"status": "success", "message": "Feedback saved"}


@app.post(
    "/api/unanswered",
    tags=["API"],
    summary="Сохранить необработанный вопрос",
    description="Сохраняет вопрос пользователя, если подходящий ответ не был найден."
)
def api_unanswered(data: UnansweredRequest):
    language = normalize_language(data.lang)
    query = data.query.strip()

    if not query:
        raise HTTPException(status_code=400, detail="Query is required")

    save_unanswered_query(query, language)
    return {"status": "success", "message": "Unanswered query saved"}


@app.get(
    "/api/question/{question_id}",
    tags=["API"],
    summary="Получить вопрос по ID",
    description="Возвращает вопрос и ответ по идентификатору вопроса и выбранному языку."
)
def api_question(question_id: int, lang: str = "ru"):
    language = normalize_language(lang)
    connection = get_db_connection()
    record = connection.execute("""
        SELECT
            Question.id,
            Category.code AS category_code,
            Translation.question_text,
            Translation.answer_text
        FROM Question
        JOIN Category ON Question.category_id = Category.id
        JOIN Translation ON Translation.question_id = Question.id
        WHERE Question.id = ? AND Translation.language = ?
    """, (question_id, language)).fetchone()
    connection.close()

    if not record:
        raise HTTPException(status_code=404, detail="Question not found")

    return {"status": "success", "question": dict(record)}



@app.get(
    "/api/health",
    tags=["API"],
    summary="Проверить состояние сервиса",
    description="Проверяет доступность приложения и подключение к базе данных PostgreSQL."
)
def api_health():
    try:
        connection = get_db_connection()
        connection.execute("SELECT 1").fetchone()
        connection.close()

        return {
            "status": "success",
            "service": "Virtual Dean",
            "database": "connected"
        }

    except Exception as error:
        raise HTTPException(
            status_code=500,
            detail=f"Database connection error: {str(error)}"
        )
    
@app.get("/", response_class=HTMLResponse, tags=["Web"])
def index(request: Request):
    language = get_request_language(request)
    return templates.TemplateResponse(
        request,
        "index.html",
        {
            "lang": language,
            "languages": LANGUAGES,
            "texts": TEXTS[language],
            "categories": get_categories(language),
            "category_groups": get_category_groups(language),
            "search_filters": get_search_filters(language)
        }
    )


@app.post("/search", response_class=HTMLResponse, tags=["Web"])
def search(
    request: Request,
    lang: str = Form("ru"),
    query: str = Form(""),
    search_filter: str = Form("all")
):
    language = normalize_language(lang)
    request.session["lang"] = language
    query = query.strip()

    if not query:
        return templates.TemplateResponse(
            request,
            "empty_query.html",
            {
                "lang": language,
                "languages": LANGUAGES,
                "texts": TEXTS[language]
            }
        )
    language_error = get_search_language_error(query, language)

    if language_error:
        return templates.TemplateResponse(
            request,
            "results.html",
            {
                "lang": language,
                "languages": LANGUAGES,
                "texts": TEXTS[language],
                "query": query,
                "main_result": None,
                "similar_results": [],
                "categories": get_categories(language),
                "search_filters": get_search_filters(language),
                "selected_filter": search_filter,
                "saved": False,
                "feedback_saved": False,
                "language_error": language_error
            }
        )
    
    main_result, similar_results = search_answers(query, language, search_filter)
    if not main_result:
        save_unanswered_query(query, language)

    return templates.TemplateResponse(
        request,
        "results.html",
        {
            "lang": language,
            "languages": LANGUAGES,
            "texts": TEXTS[language],
            "query": query,
            "main_result": main_result,
            "similar_results": similar_results,
            "categories": get_categories(language),
            "search_filters": get_search_filters(language),
            "selected_filter": search_filter,
            "saved": False,
            "feedback_saved": False,
            "language_error": ""
        }
    )


@app.post("/save-unanswered", response_class=HTMLResponse, tags=["Web"])
def save_unanswered(request: Request, lang: str = Form("ru"), query: str = Form("")):
    language = normalize_language(lang)
    request.session["lang"] = language
    query = query.strip()
    if query:
        save_unanswered_query(query, language)

    return templates.TemplateResponse(
        request,
        "results.html", 
        {

        "lang": language,
        "languages": LANGUAGES,
        "texts": TEXTS[language],
        "query": query,
        "main_result": None,
        "similar_results": [],
        "categories": get_categories(language),
        "search_filters": get_search_filters(language),
        "selected_filter": "all",
        "saved": True,
        "feedback_saved": False,
        "language_error": ""
    })


@app.post("/feedback", response_class=HTMLResponse, tags=["Web"])
def feedback(
    request: Request,
    lang: str = Form("ru"),
    query: str = Form(""),
    feedback_text: str = Form(""),
    rating: Optional[str] = Form(None)
):
    language = normalize_language(lang)
    request.session["lang"] = language
    rating_value = int(rating) if rating else None

    if feedback_text.strip() or rating_value:
        save_feedback(query.strip(), feedback_text.strip(), rating_value, language)

    return templates.TemplateResponse(
        request,
        "results.html", 
        {

        "lang": language,
        "languages": LANGUAGES,
        "texts": TEXTS[language],
        "query": query,
        "main_result": None,
        "similar_results": [],
        "categories": get_categories(language),
        "search_filters": get_search_filters(language),
        "selected_filter": "all",
        "saved": False,
        "feedback_saved": True,
        "language_error": ""
    })


@app.get("/category/{category_id}", response_class=HTMLResponse, tags=["Web"])
def category(request: Request, category_id: int):
    language = get_request_language(request)

    connection = get_db_connection()
    records = connection.execute("""
        SELECT
            Translation.question_text,
            Translation.answer_text,
            Category.code AS category_code
        FROM Translation
        JOIN Question ON Translation.question_id = Question.id
        JOIN Category ON Question.category_id = Category.id
        WHERE Translation.language = ? AND Category.id = ?
        ORDER BY Translation.question_text
    """, (language, category_id)).fetchall()
    connection.close()

    category_name = ""

    for item in get_categories(language):
        if item["id"] == category_id:
            category_name = item["name"]
            break

    return templates.TemplateResponse(
        request,
        "category.html",
        {
            "lang": language,
            "languages": LANGUAGES,
            "texts": TEXTS[language],
            "category_name": category_name,
            "records": records
        }
    )

@app.get("/admin/login", response_class=HTMLResponse, tags=["Admin"])
def admin_login_get(request: Request):
    language = get_request_language(request)
    return templates.TemplateResponse(
        request,
        "admin_login.html", 
        {

        "lang": language,
        "languages": LANGUAGES,
        "texts": TEXTS[language],
        "error": None
    })


@app.post("/admin/login", response_class=HTMLResponse, tags=["Admin"])
def admin_login_post(
    request: Request,
    username: str = Form(""),
    password: str = Form(""),
    lang: str = Form("ru")
):
    language = normalize_language(lang)
    request.session["lang"] = language

    connection = get_db_connection()
    admin = connection.execute("SELECT * FROM Admin WHERE username = ?", (username.strip(),)).fetchone()
    connection.close()

    if admin and check_password_hash(admin["password_hash"], password.strip()):
        request.session["admin_logged_in"] = True
        request.session["admin_username"] = username
        return redirect_to(f"/admin?lang={language}")

    return templates.TemplateResponse(
        request,
        "admin_login.html", 
        {

        "lang": language,
        "languages": LANGUAGES,
        "texts": TEXTS[language],
        "error": "Неверный логин или пароль"
    })


@app.get("/admin", tags=["Admin"])
def admin_panel(request: Request):
    language = get_request_language(request)
    if not is_admin_logged_in(request):
        return redirect_to(f"/admin/login?lang={language}")
    return redirect_to(f"/admin/unanswered?lang={language}")


def get_current_admin_id(request: Request):
    username = request.session.get("admin_username")

    if not username:
        return None

    connection = get_db_connection()

    admin = connection.execute("""
        SELECT id
        FROM Admin
        WHERE username = ?
    """, (username,)).fetchone()

    connection.close()

    if not admin:
        return None

    return admin["id"]
@app.get("/admin/unanswered", response_class=HTMLResponse, tags=["Admin"])
def admin_unanswered(request: Request):
    language = get_request_language(request)
    if not is_admin_logged_in(request):
        return redirect_to(f"/admin/login?lang={language}")

    connection = get_db_connection()
    unanswered_queries = connection.execute("SELECT * FROM UnansweredQuery ORDER BY created_at DESC").fetchall()
    unanswered_by_language = connection.execute("""
        SELECT language, COUNT(*) AS count
        FROM UnansweredQuery
        GROUP BY language
        ORDER BY count DESC
    """).fetchall()
    unanswered_by_date = connection.execute("""
        SELECT TO_CHAR(created_at, 'YYYY-MM-DD') AS date, COUNT(*) AS count
        FROM UnansweredQuery
        GROUP BY TO_CHAR(created_at, 'YYYY-MM-DD')
        ORDER BY date DESC
        LIMIT 7
    """).fetchall()
    connection.close()

    return templates.TemplateResponse(
        request,
        "admin_unanswered.html", 
        {
        
        "lang": language,
        "languages": LANGUAGES,
        "texts": TEXTS[language],
        "unanswered_queries": unanswered_queries,
        "unanswered_by_language": unanswered_by_language,
        "unanswered_by_date": unanswered_by_date
    })



@app.get(
    "/admin/unanswered/add/{item_id}",
    response_class=HTMLResponse,
    tags=["Admin"]
)
def admin_add_from_unanswered_get(request: Request, item_id: int, error: str = ""):
    language = get_request_language(request)

    if not is_admin_logged_in(request):
        return redirect_to(f"/admin/login?lang={language}")

    connection = get_db_connection()

    unanswered_item = connection.execute("""
        SELECT *
        FROM UnansweredQuery
        WHERE id = ?
    """, (item_id,)).fetchone()

    categories = connection.execute("""
        SELECT *
        FROM Category
        ORDER BY code
    """).fetchall()

    connection.close()

    if not unanswered_item:
        raise HTTPException(status_code=404, detail="Unanswered query not found")

    return templates.TemplateResponse(
        request,
        "admin_add_from_unanswered.html",
        {
            "lang": language,
            "languages": LANGUAGES,
            "texts": TEXTS[language],
            "item": unanswered_item,
            "categories": categories,
            "error": error,
            "category_names": CATEGORY_NAMES[language]
        }
    )

@app.post(
    "/admin/unanswered/add/{item_id}",
    tags=["Admin"]
)
def admin_add_from_unanswered_post(
    request: Request,
    item_id: int,
    category_id: str = Form(...),
    question_ru: str = Form(""),
    answer_ru: str = Form(""),
    question_en: str = Form(""),
    answer_en: str = Form(""),
    question_zh: str = Form(""),
    answer_zh: str = Form("")
):
    language = get_request_language(request)

    if not is_admin_logged_in(request):
        return redirect_to(f"/admin/login?lang={language}")

    if (
        not question_ru.strip()
        or not answer_ru.strip()
        or not question_en.strip()
        or not answer_en.strip()
        or not question_zh.strip()
        or not answer_zh.strip()
    ):
        error_text = quote("Заполните все поля: русский, английский и китайский вопрос и ответ.")
        return redirect_to(
            f"/admin/unanswered/add/{item_id}?lang={language}&error={error_text}"
        )



    translations = [
        ("ru", question_ru.strip(), answer_ru.strip()),
        ("en", question_en.strip(), answer_en.strip()),
        ("zh", question_zh.strip(), answer_zh.strip())
    ]

    language_errors = validate_question_languages(translations)

    if language_errors:
        error_text = quote(" ".join(language_errors))
        return redirect_to(
            f"/admin/unanswered/add/{item_id}?lang={language}&error={error_text}"
        )

    admin_id = get_current_admin_id(request)

    connection = get_db_connection()

    cursor = connection.execute("""
        INSERT INTO Question (category_id, created_by_admin_id, updated_by_admin_id)
        VALUES (?, ?, ?)
        RETURNING id
    """, (category_id, admin_id, admin_id))
    question_id = cursor.fetchone()["id"]

    for lang_code, question_text, answer_text in translations:
        connection.execute("""
            INSERT INTO Translation (question_id, language, question_text, answer_text)
            VALUES (?, ?, ?, ?)
            ON CONFLICT (question_id, language) DO NOTHING
        """, (question_id, lang_code, question_text, answer_text))
        
    connection.execute("""
        UPDATE UnansweredQuery
        SET status = ?,
            processed_question_id = ?,
            processed_by_admin_id = ?
        WHERE id = ?
    """, ("processed", question_id, admin_id, item_id))

    connection.commit()
    connection.close()

    return redirect_to(f"/admin/unanswered?lang={language}")

@app.post("/admin/unanswered/delete/{item_id}", tags=["Admin"])
def admin_delete_unanswered(request: Request, item_id: int):
    language = get_request_language(request)
    if not is_admin_logged_in(request):
        return redirect_to(f"/admin/login?lang={language}")
    connection = get_db_connection()
    connection.execute("DELETE FROM UnansweredQuery WHERE id = ?", (item_id,))
    connection.commit()
    connection.close()
    return redirect_to(f"/admin/unanswered?lang={language}")


@app.post("/admin/unanswered/clear", tags=["Admin"])
def admin_clear_unanswered(request: Request):
    language = get_request_language(request)
    if not is_admin_logged_in(request):
        return redirect_to(f"/admin/login?lang={language}")
    connection = get_db_connection()
    connection.execute("DELETE FROM UnansweredQuery")
    connection.commit()
    connection.close()
    return redirect_to(f"/admin/unanswered?lang={language}")


@app.get("/admin/feedback", response_class=HTMLResponse, tags=["Admin"])
def admin_feedback(request: Request):
    language = get_request_language(request)
    if not is_admin_logged_in(request):
        return redirect_to(f"/admin/login?lang={language}")

    connection = get_db_connection()
    feedback_items = connection.execute("SELECT * FROM Feedback ORDER BY created_at DESC").fetchall()
    rating_stats = connection.execute("""
        SELECT rating, COUNT(*) AS count
        FROM Feedback
        WHERE rating IS NOT NULL
        GROUP BY rating
        ORDER BY rating DESC
    """).fetchall()
    avg_rating = connection.execute("""
        SELECT ROUND(AVG(rating), 2) AS value
        FROM Feedback
        WHERE rating IS NOT NULL
    """).fetchone()
    feedback_by_language = connection.execute("""
        SELECT language, COUNT(*) AS count
        FROM Feedback
        GROUP BY language
        ORDER BY count DESC
    """).fetchall()
    feedback_by_date = connection.execute("""
        SELECT TO_CHAR(created_at, 'YYYY-MM-DD') AS date, COUNT(*) AS count
        FROM Feedback
        GROUP BY TO_CHAR(created_at, 'YYYY-MM-DD')
        ORDER BY date DESC
        LIMIT 7
    """).fetchall()
    connection.close()

    return templates.TemplateResponse(
        request,

        "admin_feedback.html", 
        {

        "lang": language,
        "languages": LANGUAGES,
        "texts": TEXTS[language],
        "feedback_items": feedback_items,
        "rating_stats": rating_stats,
        "avg_rating": avg_rating,
        "feedback_by_language": feedback_by_language,
        "feedback_by_date": feedback_by_date
    })


@app.post("/admin/feedback/delete/{item_id}", tags=["Admin"])
def admin_delete_feedback(request: Request, item_id: int):
    language = get_request_language(request)
    if not is_admin_logged_in(request):
        return redirect_to(f"/admin/login?lang={language}")
    connection = get_db_connection()
    connection.execute("DELETE FROM Feedback WHERE id = ?", (item_id,))
    connection.commit()
    connection.close()
    return redirect_to(f"/admin/feedback?lang={language}")


@app.post("/admin/feedback/clear", tags=["Admin"])
def admin_clear_feedback(request: Request):
    language = get_request_language(request)
    if not is_admin_logged_in(request):
        return redirect_to(f"/admin/login?lang={language}")
    connection = get_db_connection()
    connection.execute("DELETE FROM Feedback")
    connection.commit()
    connection.close()
    return redirect_to(f"/admin/feedback?lang={language}")


@app.get("/admin/knowledge", response_class=HTMLResponse, tags=["Admin"])
def admin_knowledge(
    request: Request,
    category: str = "all",
    q: str = "",
    page: int = 1,
    error: str = ""
):
    language = get_request_language(request)
    if not is_admin_logged_in(request):
        return redirect_to(f"/admin/login?lang={language}")

    selected_category = category
    search_query = q.strip()
    per_page = 20
    offset = max(page - 1, 0) * per_page

    connection = get_db_connection()
    categories = connection.execute("SELECT * FROM Category ORDER BY code").fetchall()

    where_conditions = []
    params = []
    if selected_category != "all":
        where_conditions.append("Category.code = ?")
        params.append(selected_category)
    if search_query:
        where_conditions.append("Translation.question_text ILIKE ?")
        params.append(f"%{search_query}%")

    where_sql = ""
    if where_conditions:
        where_sql = "WHERE " + " AND ".join(where_conditions)

    total_row = connection.execute(f"""
        SELECT COUNT(*) AS total
        FROM Question
        JOIN Category ON Question.category_id = Category.id
        LEFT JOIN Translation ON Translation.question_id = Question.id
            AND Translation.language = 'ru'
        {where_sql}
    """, params).fetchone()

    total_questions = total_row["total"]
    total_pages = (total_questions + per_page - 1) // per_page

    questions = connection.execute(f"""
        SELECT
            Question.id,
            Category.code AS category_code,
            Translation.question_text
        FROM Question
        JOIN Category ON Question.category_id = Category.id
        LEFT JOIN Translation ON Translation.question_id = Question.id
            AND Translation.language = 'ru'
        {where_sql}
        ORDER BY Question.id DESC
        LIMIT ? OFFSET ?
    """, params + [per_page, offset]).fetchall()
    connection.close()

    return templates.TemplateResponse(
        request,
        
        "admin_knowledge.html", 
        {
        "lang": language,
        "error": error,
        "languages": LANGUAGES,
        "texts": TEXTS[language],
        "questions": questions,
        "categories": categories,
        "category_names": CATEGORY_NAMES[language],
        "selected_category": selected_category,
        "search_query": search_query,
        "page": page,
        "total_pages": total_pages,
        "total_questions": total_questions
    })


@app.post("/admin/knowledge/add", tags=["Admin"])
def admin_add_question(
    request: Request,
    category_id: str = Form(...),
    question_ru: str = Form(""),
    answer_ru: str = Form(""),
    question_en: str = Form(""),
    answer_en: str = Form(""),
    question_zh: str = Form(""),
    answer_zh: str = Form("")
):
    language = get_request_language(request)

    if not is_admin_logged_in(request):
        return redirect_to(f"/admin/login?lang={language}")

    if (
        not category_id
        or not question_ru.strip()
        or not answer_ru.strip()
        or not question_en.strip()
        or not answer_en.strip()
        or not question_zh.strip()
        or not answer_zh.strip()
    ):
        error_text = quote("Заполните все поля: русский, английский и китайский вопрос и ответ.")
        return redirect_to(f"/admin/knowledge?lang={language}&error={error_text}")

    translations = [
        ("ru", question_ru.strip(), answer_ru.strip()),
        ("en", question_en.strip(), answer_en.strip()),
        ("zh", question_zh.strip(), answer_zh.strip())
    ]

    language_errors = validate_question_languages(translations)

    if language_errors:
        error_text = quote(" ".join(language_errors))
        return redirect_to(f"/admin/knowledge?lang={language}&error={error_text}")

    admin_id = get_current_admin_id(request)

    connection = get_db_connection()

    cursor = connection.execute("""
        INSERT INTO Question (category_id, created_by_admin_id, updated_by_admin_id)
        VALUES (?, ?, ?)
        RETURNING id
    """, (category_id, admin_id, admin_id))

    question_id = cursor.fetchone()["id"]

    for lang_code, question_text, answer_text in translations:
        connection.execute("""
            INSERT INTO Translation (question_id, language, question_text, answer_text)
            VALUES (?, ?, ?, ?)
            ON CONFLICT (question_id, language) DO NOTHING
        """, (question_id, lang_code, question_text, answer_text))

    connection.commit()
    connection.close()

    return redirect_to(f"/admin/knowledge?lang={language}")



@app.post("/admin/knowledge/delete/{question_id}", tags=["Admin"])
def admin_delete_question(request: Request, question_id: int):
    language = get_request_language(request)
    if not is_admin_logged_in(request):
        return redirect_to(f"/admin/login?lang={language}")
    connection = get_db_connection()
    connection.execute("DELETE FROM Translation WHERE question_id = ?", (question_id,))
    connection.execute("DELETE FROM Question WHERE id = ?", (question_id,))
    connection.commit()
    connection.close()
    return redirect_to(f"/admin/knowledge?lang={language}")


@app.get("/admin/knowledge/edit/{question_id}", response_class=HTMLResponse, tags=["Admin"])
def admin_edit_question_get(request: Request, question_id: int, error: str = ""):
    language = get_request_language(request)

    if not is_admin_logged_in(request):
        return redirect_to(f"/admin/login?lang={language}")

    connection = get_db_connection()

    question = connection.execute("""
        SELECT Question.id, Question.category_id, Category.code AS category_code
        FROM Question
        JOIN Category ON Question.category_id = Category.id
        WHERE Question.id = ?
    """, (question_id,)).fetchone()

    if not question:
        connection.close()
        raise HTTPException(status_code=404, detail="Question not found")

    translation_rows = connection.execute("""
        SELECT language, question_text, answer_text
        FROM Translation
        WHERE question_id = ?
    """, (question_id,)).fetchall()

    categories = connection.execute("""
        SELECT *
        FROM Category
        ORDER BY code
    """).fetchall()

    connection.close()

    translations = {
        "ru": {"question": "", "answer": ""},
        "en": {"question": "", "answer": ""},
        "zh": {"question": "", "answer": ""}
    }

    for row in translation_rows:
        translations[row["language"]] = {
            "question": row["question_text"],
            "answer": row["answer_text"]
        }

    return templates.TemplateResponse(
        request,
        "admin_edit_question.html",
        {
            "lang": language,
            "error": error,
            "languages": LANGUAGES,
            "texts": TEXTS[language],
            "question": question,
            "translations": translations,
            "categories": categories,
            "category_names": CATEGORY_NAMES[language]
        }
    )


@app.post("/admin/knowledge/edit/{question_id}", tags=["Admin"])
def admin_edit_question_post(
    request: Request,
    question_id: int,
    category_id: str = Form(...),
    question_ru: str = Form(""),
    answer_ru: str = Form(""),
    question_en: str = Form(""),
    answer_en: str = Form(""),
    question_zh: str = Form(""),
    answer_zh: str = Form("")
):
    language = get_request_language(request)

    if not is_admin_logged_in(request):
        return redirect_to(f"/admin/login?lang={language}")

    if (
        not question_ru.strip()
        or not answer_ru.strip()
        or not question_en.strip()
        or not answer_en.strip()
        or not question_zh.strip()
        or not answer_zh.strip()
    ):
        error_text = quote("Заполните все поля: русский, английский и китайский вопрос и ответ.")
        return redirect_to(
            f"/admin/knowledge/edit/{question_id}?lang={language}&error={error_text}"
        )

    translations = [
        ("ru", question_ru.strip(), answer_ru.strip()),
        ("en", question_en.strip(), answer_en.strip()),
        ("zh", question_zh.strip(), answer_zh.strip())
    ]

    language_errors = validate_question_languages(translations)

    if language_errors:
        error_text = quote(" ".join(language_errors))
        return redirect_to(
            f"/admin/knowledge/edit/{question_id}?lang={language}&error={error_text}"
        )
    admin_id = get_current_admin_id(request)

    connection = get_db_connection()

    connection.execute("""
        UPDATE Question
        SET category_id = ?,
            updated_by_admin_id = ?
        WHERE id = ?
    """, (category_id, admin_id, question_id))

    for lang_code, question_text, answer_text in translations:
        existing = connection.execute("""
            SELECT id
            FROM Translation
            WHERE question_id = ? AND language = ?
        """, (question_id, lang_code)).fetchone()

        if existing:
            connection.execute("""
                UPDATE Translation
                SET question_text = ?, answer_text = ?
                WHERE question_id = ? AND language = ?
            """, (question_text, answer_text, question_id, lang_code))
        else:
            connection.execute("""
                INSERT INTO Translation (question_id, language, question_text, answer_text)
                VALUES (?, ?, ?, ?)
            """, (question_id, lang_code, question_text, answer_text))

    connection.commit()
    connection.close()

    return redirect_to(f"/admin/knowledge?lang={language}")

@app.get("/admin/unanswered/export", tags=["Admin"])
def export_unanswered_excel(request: Request):
    language = get_request_language(request)
    if not is_admin_logged_in(request):
        return redirect_to(f"/admin/login?lang={language}")

    connection = get_db_connection()
    rows = connection.execute("""
        SELECT id, query_text, language, created_at, status
        FROM UnansweredQuery
        ORDER BY created_at DESC
    """).fetchall()
    by_language = connection.execute("""
        SELECT language, COUNT(*) AS count
        FROM UnansweredQuery
        GROUP BY language
        ORDER BY count DESC
    """).fetchall()
    by_date = connection.execute("""
        SELECT TO_CHAR(created_at, 'YYYY-MM-DD') AS date, COUNT(*) AS count
        FROM UnansweredQuery
        GROUP BY TO_CHAR(created_at, 'YYYY-MM-DD')
        ORDER BY date DESC
    """).fetchall()
    connection.close()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Запросы"
    sheet.append(["ID", "Запрос", "Язык", "Дата", "Статус"])
    for row in rows:
        sheet.append([row["id"], row["query_text"], row["language"], row["created_at"], row["status"]])

    stats_sheet = workbook.create_sheet("Аналитика")
    stats_sheet.append(["По языкам"])
    stats_sheet.append(["Язык", "Количество"])
    for item in by_language:
        stats_sheet.append([item["language"], item["count"]])
    stats_sheet.append([])
    stats_sheet.append(["По датам"])
    stats_sheet.append(["Дата", "Количество"])
    for item in by_date:
        stats_sheet.append([item["date"], item["count"]])

    format_excel_workbook(workbook)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=unanswered_queries_report.xlsx"}
    )


@app.get("/admin/feedback/export", tags=["Admin"])
def export_feedback_excel(request: Request):
    language = get_request_language(request)
    if not is_admin_logged_in(request):
        return redirect_to(f"/admin/login?lang={language}")

    connection = get_db_connection()
    rows = connection.execute("""
        SELECT id, query_text, feedback_text, rating, language, created_at, status
        FROM Feedback
        ORDER BY created_at DESC
    """).fetchall()
    rating_stats = connection.execute("""
        SELECT rating, COUNT(*) AS count
        FROM Feedback
        WHERE rating IS NOT NULL
        GROUP BY rating
        ORDER BY rating DESC
    """).fetchall()
    avg_rating = connection.execute("""
        SELECT ROUND(AVG(rating), 2) AS value
        FROM Feedback
        WHERE rating IS NOT NULL
    """).fetchone()
    by_language = connection.execute("""
        SELECT language, COUNT(*) AS count
        FROM Feedback
        GROUP BY language
        ORDER BY count DESC
    """).fetchall()
    by_date = connection.execute("""
        SELECT TO_CHAR(created_at, 'YYYY-MM-DD') AS date, COUNT(*) AS count
        FROM Feedback
        GROUP BY TO_CHAR(created_at, 'YYYY-MM-DD')
        ORDER BY date DESC
    """).fetchall()
    connection.close()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Обратная связь"
    sheet.append(["ID", "Запрос", "Сообщение", "Оценка", "Язык", "Дата", "Статус"])
    for row in rows:
        sheet.append([row["id"], row["query_text"], row["feedback_text"], row["rating"], row["language"], row["created_at"], row["status"]])

    stats_sheet = workbook.create_sheet("Аналитика")
    stats_sheet.append(["Средняя оценка", avg_rating["value"] if avg_rating and avg_rating["value"] else "—"])
    stats_sheet.append([])
    stats_sheet.append(["Оценки"])
    stats_sheet.append(["Оценка", "Количество"])
    for item in rating_stats:
        stats_sheet.append([item["rating"], item["count"]])
    stats_sheet.append([])
    stats_sheet.append(["По языкам"])
    stats_sheet.append(["Язык", "Количество"])
    for item in by_language:
        stats_sheet.append([item["language"], item["count"]])
    stats_sheet.append([])
    stats_sheet.append(["По датам"])
    stats_sheet.append(["Дата", "Количество"])
    for item in by_date:
        stats_sheet.append([item["date"], item["count"]])

    format_excel_workbook(workbook)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=feedback_report.xlsx"}
    )


@app.get("/admin/logout", tags=["Admin"])
def admin_logout(request: Request):
    language = get_request_language(request)
    request.session.pop("admin_logged_in", None)
    request.session.pop("admin_username", None)
    return redirect_to(f"/?lang={language}")
